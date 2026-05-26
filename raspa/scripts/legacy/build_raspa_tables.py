#!/usr/bin/env python3
import csv
import json
import math
import re
import sys
from pathlib import Path
from typing import Dict, List, Optional, Tuple

MOLAR_MASS_H2 = 2.01588
ROOT = Path.home() / "raspa_runs" / "H2_in_MOF"
DATASETS_DIR = Path.home() / "datasets"
CASES = ["PS_high", "PS_low", "TPS_low"]

OUT_SUMMARY = ROOT / "ALL_corrected_summary_with_sanity.csv"
OUT_ABS_EXCESS = ROOT / "ALL_corrected_summary_absolute_vs_excess.csv"
OUT_JOIN_PS = ROOT / "MINE_vs_DATASETS_PS.csv"
OUT_JOIN_TPS = ROOT / "MINE_vs_DATASETS_TPS.csv"


def clean_text(s: Optional[str]) -> str:
    if s is None:
        return ""
    return str(s).replace("\xa0", " ").strip()


def to_float(s: Optional[str]) -> Optional[float]:
    if s is None:
        return None
    t = clean_text(s)
    if not t:
        return None
    t = t.replace(",", "")
    try:
        return float(t)
    except ValueError:
        return None


def fmt(v: Optional[float], nd: int = 6) -> str:
    if v is None:
        return ""
    return f"{v:.{nd}f}"


def norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_text(h).lower())


def norm_id(s: str) -> str:
    t = clean_text(s).upper()
    t = re.sub(r"\s+", "_", t)
    t = t.replace("-", "_")
    t = re.sub(r"[^A-Z0-9_]", "", t)
    t = re.sub(r"(_(CLEAN|H2|CORE))+$", "", t)
    t = re.sub(r"_+$", "", t)
    return t


def id_candidates(s: str) -> List[str]:
    base = norm_id(s)
    out = set()
    if base:
        out.add(base)
        out.add(re.sub(r"_(CLEAN|H2|CORE)$", "", base))
        out.add(re.sub(r"_(CLEAN|H2|CORE)$", "", re.sub(r"_(CLEAN|H2|CORE)$", "", base)))
        if "_" in base:
            out.add(base.split("_")[0])
        out.add(re.sub(r"[^A-Z0-9]", "", base))
    return [x for x in out if x]


def detect_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=[",", "\t", ";", "|"])
        return dialect.delimiter
    except Exception:
        if sample.count("\t") > sample.count(","):
            return "\t"
        return ","


def read_text(path: Path) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return path.read_text(encoding=enc)
        except Exception:
            continue
    return path.read_text(errors="ignore")


def load_table(path: Path) -> Tuple[List[str], List[Dict[str, str]], str]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        try:
            import openpyxl  # type: ignore
        except Exception as exc:
            raise RuntimeError(f"Cannot read {path} (xlsx) without openpyxl: {exc}")
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return [], [], "xlsx"
        headers = [clean_text(x) for x in rows[0]]
        data = []
        for r in rows[1:]:
            rec = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                v = r[i] if i < len(r) else ""
                rec[h] = "" if v is None else str(v)
            data.append(rec)
        return headers, data, "xlsx"

    if suffix == ".json":
        obj = json.loads(read_text(path))
        if isinstance(obj, list):
            rows = []
            keys = set()
            for row in obj:
                if isinstance(row, dict):
                    keys.update(row.keys())
            headers = sorted(keys)
            for row in obj:
                if not isinstance(row, dict):
                    continue
                rec = {k: clean_text(row.get(k, "")) for k in headers}
                rows.append(rec)
            return headers, rows, "json"
        raise RuntimeError(f"Unsupported JSON structure in {path}")

    text = read_text(path)
    delim = detect_delimiter("\n".join(text.splitlines()[:50]))
    lines = text.splitlines()
    if not lines:
        return [], [], f"delim:{delim}"

    reader = csv.DictReader(lines, delimiter=delim)
    if reader.fieldnames is None:
        return [], [], f"delim:{delim}"
    headers = [clean_text(h) for h in reader.fieldnames]

    rows = []
    for raw in reader:
        rec = {}
        for k, v in raw.items():
            if k is None:
                continue
            ck = clean_text(k)
            rec[ck] = clean_text(v)
        rows.append(rec)
    return headers, rows, f"delim:{delim}"


def extract_last_float(pattern: str, text: str) -> Optional[float]:
    m = re.findall(pattern, text, flags=re.MULTILINE)
    if not m:
        return None
    return to_float(m[-1])


def extract_first_float(pattern: str, text: str) -> Optional[float]:
    m = re.search(pattern, text, flags=re.MULTILINE)
    if not m:
        return None
    return to_float(m.group(1))


def parse_output(output_file: Path) -> Dict[str, Optional[float]]:
    text = read_text(output_file)

    abs_molkg = extract_last_float(r"Average loading absolute \[mol/kg framework\]\s+([-+0-9.eE]+)", text)
    abs_mgg = extract_last_float(r"Average loading absolute \[milligram/gram framework\]\s+([-+0-9.eE]+)", text)
    excess_molkg = extract_last_float(r"Average loading excess \[mol/kg framework\]\s+([-+0-9.eE]+)", text)
    excess_mgg = extract_last_float(r"Average loading excess \[milligram/gram framework\]\s+([-+0-9.eE]+)", text)

    density = extract_first_float(r"Framework Density:\s+([-+0-9.eE]+)\s+\[kg/m\^3\]", text)
    fugacity = extract_first_float(r"Fugacity coefficient:\s+([-+0-9.eE]+)", text)

    token = ""
    m = re.search(r"atom:\s+\d+\s+is of type:[^\[]*\[\s*([A-Za-z0-9_+\-]+)\s*\]", text)
    if m:
        token = clean_text(m.group(1))

    abs_mgg_corr = abs_molkg * MOLAR_MASS_H2 if abs_molkg is not None else None
    excess_mgg_corr = excess_molkg * MOLAR_MASS_H2 if excess_molkg is not None else None

    abs_ratio = None
    if abs_mgg is not None and abs_mgg_corr not in (None, 0.0):
        abs_ratio = abs_mgg / abs_mgg_corr

    excess_ratio = None
    if excess_mgg is not None and excess_mgg_corr not in (None, 0.0):
        excess_ratio = excess_mgg / excess_mgg_corr

    abs_wt = abs_mgg_corr / 10.0 if abs_mgg_corr is not None else None
    abs_uv = None
    if abs_molkg is not None and density is not None:
        abs_uv = abs_molkg * density * MOLAR_MASS_H2 / 1000.0

    excess_wt = excess_mgg_corr / 10.0 if excess_mgg_corr is not None else None
    excess_uv = None
    if excess_molkg is not None and density is not None:
        excess_uv = excess_molkg * density * MOLAR_MASS_H2 / 1000.0

    return {
        "abs_molkg": abs_molkg,
        "abs_mgg_reported": abs_mgg,
        "abs_mgg_corrected": abs_mgg_corr,
        "ratio_reported": abs_ratio,
        "framework_density_kg_m3": density,
        "abs_wt_pct": abs_wt,
        "abs_uv_gL": abs_uv,
        "excess_molkg": excess_molkg,
        "excess_mgg_reported": excess_mgg,
        "excess_mgg_corrected": excess_mgg_corr,
        "excess_ratio_reported": excess_ratio,
        "excess_wt_pct": excess_wt,
        "excess_uv_gL": excess_uv,
        "fugacity_coefficient": fugacity,
        "pseudo_atom_token": token,
    }


def parse_simulation_input(sim_file: Path) -> Dict[str, str]:
    out = {
        "forcefield": "",
        "moleculedefinitions": "",
        "component_moleculedefinition": "",
        "eos_lines": "",
    }
    if not sim_file.exists():
        return out

    eos = []
    for line in read_text(sim_file).splitlines():
        s = line.strip()
        if not s:
            continue
        parts = s.split()
        if not parts:
            continue
        key = parts[0]
        if key == "ForceField" and len(parts) >= 2:
            out["forcefield"] = parts[1]
        elif key == "MoleculeDefinitions" and len(parts) >= 2:
            out["moleculedefinitions"] = parts[1]
        elif key == "MoleculeDefinition" and len(parts) >= 2:
            out["component_moleculedefinition"] = parts[1]
        if any(k in s for k in ["UseFugacityCoefficients", "FugacityCoefficient", "EquationOfState"]):
            eos.append(s)
    out["eos_lines"] = "; ".join(eos)
    return out


def find_newest_output(mof_dir: Path, case: str) -> Optional[Path]:
    cands: List[Path] = []
    for p in mof_dir.rglob("output_*.data"):
        sp = p.as_posix()
        if f"/{case}/" in sp and "/Output/System_0/" in sp:
            cands.append(p)
    if not cands:
        return None
    return max(cands, key=lambda x: x.stat().st_mtime)


def safe_sub(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if a is None or b is None:
        return None
    return a - b


def choose_best_column(headers: List[str], include: List[str], exclude: Optional[List[str]] = None) -> Tuple[str, List[Tuple[str, int]]]:
    if exclude is None:
        exclude = []
    scored = []
    for h in headers:
        nh = norm_header(h)
        score = 0
        if any(x in nh for x in exclude):
            score -= 2
        for tok in include:
            if tok in nh:
                score += 3
        if all(tok in nh for tok in include):
            score += 4
        scored.append((h, score))
    scored.sort(key=lambda x: (-x[1], x[0]))
    best = scored[0][0] if scored else ""
    return best, scored[:8]


def build_dataset_index(rows: List[Dict[str, str]], key_cols: List[str]) -> Dict[str, Dict[str, str]]:
    idx: Dict[str, Dict[str, str]] = {}
    for row in rows:
        for col in key_cols:
            if not col:
                continue
            raw = row.get(col, "")
            for cand in id_candidates(raw):
                idx.setdefault(cand, row)
    return idx


def pct_diff(mine: Optional[float], ref: Optional[float]) -> Optional[float]:
    if mine is None or ref is None or ref == 0:
        return None
    return 100.0 * (mine - ref) / ref


def likely_reason(
    mine_ug: Optional[float],
    mine_uv: Optional[float],
    ref_ug: Optional[float],
    ref_uv: Optional[float],
    dataset_excess_hint: bool,
) -> str:
    if ref_ug is None and ref_uv is None:
        return "no dataset match"

    ug_pct = abs(pct_diff(mine_ug, ref_ug)) if pct_diff(mine_ug, ref_ug) is not None else None
    uv_pct = abs(pct_diff(mine_uv, ref_uv)) if pct_diff(mine_uv, ref_uv) is not None else None

    if dataset_excess_hint and mine_ug is not None and ref_ug is not None and mine_ug > ref_ug * 1.05:
        return "absolute vs excess likely"
    if ug_pct is not None and uv_pct is not None and uv_pct > ug_pct + 20.0:
        return "density basis likely (crystal vs bulk)"
    if (ug_pct is not None and ug_pct > 20.0) or (uv_pct is not None and uv_pct > 20.0):
        return "isotherm curvature / sampling difference"
    return "minor mismatch / likely comparable basis"


def main() -> int:
    mof_dirs = sorted([p for p in ROOT.glob("*_H2") if p.is_dir()])
    mine_rows = []
    has_excess_any = False

    for mof_dir in mof_dirs:
        mof_name = mof_dir.name[:-3] if mof_dir.name.endswith("_H2") else mof_dir.name
        row: Dict[str, str] = {"MOF": mof_name}
        case_data: Dict[str, Dict[str, Optional[float]]] = {}

        for case in CASES:
            out = find_newest_output(mof_dir, case)
            prefix = f"{case}_"
            if out is None:
                for k in [
                    "output_data", "simulation_input", "run_sh", "forcefield", "moleculedefinitions",
                    "component_moleculedefinition", "eos_lines", "pseudo_atom_token", "abs_molkg",
                    "abs_mgg_reported", "abs_mgg_corrected", "ratio_reported", "framework_density_kg_m3",
                    "abs_wt_pct", "abs_uv_gL", "excess_molkg", "excess_mgg_reported",
                    "excess_mgg_corrected", "excess_ratio_reported", "excess_wt_pct", "excess_uv_gL",
                    "fugacity_coefficient",
                ]:
                    row[prefix + k] = ""
                continue

            case_dir = out.parents[2]
            sim_file = case_dir / "simulation.input"
            run_file = case_dir / "run.sh"

            row[prefix + "output_data"] = str(out)
            row[prefix + "simulation_input"] = str(sim_file)
            row[prefix + "run_sh"] = str(run_file)

            sim = parse_simulation_input(sim_file)
            row[prefix + "forcefield"] = sim["forcefield"]
            row[prefix + "moleculedefinitions"] = sim["moleculedefinitions"]
            row[prefix + "component_moleculedefinition"] = sim["component_moleculedefinition"]
            row[prefix + "eos_lines"] = sim["eos_lines"]

            parsed = parse_output(out)
            case_data[case] = parsed
            if parsed.get("excess_molkg") is not None or parsed.get("excess_mgg_reported") is not None:
                has_excess_any = True

            row[prefix + "pseudo_atom_token"] = str(parsed.get("pseudo_atom_token") or "")
            row[prefix + "abs_molkg"] = fmt(parsed.get("abs_molkg"), 10)
            row[prefix + "abs_mgg_reported"] = fmt(parsed.get("abs_mgg_reported"), 10)
            row[prefix + "abs_mgg_corrected"] = fmt(parsed.get("abs_mgg_corrected"), 10)
            row[prefix + "ratio_reported"] = fmt(parsed.get("ratio_reported"), 6)
            row[prefix + "framework_density_kg_m3"] = fmt(parsed.get("framework_density_kg_m3"), 12)
            row[prefix + "abs_wt_pct"] = fmt(parsed.get("abs_wt_pct"), 6)
            row[prefix + "abs_uv_gL"] = fmt(parsed.get("abs_uv_gL"), 6)
            row[prefix + "excess_molkg"] = fmt(parsed.get("excess_molkg"), 10)
            row[prefix + "excess_mgg_reported"] = fmt(parsed.get("excess_mgg_reported"), 10)
            row[prefix + "excess_mgg_corrected"] = fmt(parsed.get("excess_mgg_corrected"), 10)
            row[prefix + "excess_ratio_reported"] = fmt(parsed.get("excess_ratio_reported"), 6)
            row[prefix + "excess_wt_pct"] = fmt(parsed.get("excess_wt_pct"), 6)
            row[prefix + "excess_uv_gL"] = fmt(parsed.get("excess_uv_gL"), 6)
            row[prefix + "fugacity_coefficient"] = fmt(parsed.get("fugacity_coefficient"), 10)

        abs_ug_ps = safe_sub(case_data.get("PS_high", {}).get("abs_wt_pct"), case_data.get("PS_low", {}).get("abs_wt_pct"))
        abs_uv_ps = safe_sub(case_data.get("PS_high", {}).get("abs_uv_gL"), case_data.get("PS_low", {}).get("abs_uv_gL"))
        abs_ug_tps = safe_sub(case_data.get("PS_high", {}).get("abs_wt_pct"), case_data.get("TPS_low", {}).get("abs_wt_pct"))
        abs_uv_tps = safe_sub(case_data.get("PS_high", {}).get("abs_uv_gL"), case_data.get("TPS_low", {}).get("abs_uv_gL"))

        exc_ug_ps = safe_sub(case_data.get("PS_high", {}).get("excess_wt_pct"), case_data.get("PS_low", {}).get("excess_wt_pct"))
        exc_uv_ps = safe_sub(case_data.get("PS_high", {}).get("excess_uv_gL"), case_data.get("PS_low", {}).get("excess_uv_gL"))
        exc_ug_tps = safe_sub(case_data.get("PS_high", {}).get("excess_wt_pct"), case_data.get("TPS_low", {}).get("excess_wt_pct"))
        exc_uv_tps = safe_sub(case_data.get("PS_high", {}).get("excess_uv_gL"), case_data.get("TPS_low", {}).get("excess_uv_gL"))

        row["UG_PS"] = fmt(abs_ug_ps, 6)
        row["UV_PS"] = fmt(abs_uv_ps, 6)
        row["UG_TPS"] = fmt(abs_ug_tps, 6)
        row["UV_TPS"] = fmt(abs_uv_tps, 6)

        row["UG_PS_excess"] = fmt(exc_ug_ps, 6)
        row["UV_PS_excess"] = fmt(exc_uv_ps, 6)
        row["UG_TPS_excess"] = fmt(exc_ug_tps, 6)
        row["UV_TPS_excess"] = fmt(exc_uv_tps, 6)

        mine_rows.append(row)

    summary_fields = ["MOF"]
    per_case_suffixes = [
        "output_data", "simulation_input", "run_sh", "forcefield", "moleculedefinitions",
        "component_moleculedefinition", "eos_lines", "pseudo_atom_token", "abs_molkg",
        "abs_mgg_reported", "abs_mgg_corrected", "ratio_reported", "framework_density_kg_m3",
        "abs_wt_pct", "abs_uv_gL", "excess_molkg", "excess_mgg_reported", "excess_mgg_corrected",
        "excess_ratio_reported", "excess_wt_pct", "excess_uv_gL", "fugacity_coefficient",
    ]
    for case in CASES:
        summary_fields.extend([f"{case}_{s}" for s in per_case_suffixes])
    summary_fields.extend(["UG_PS", "UV_PS", "UG_TPS", "UV_TPS", "UG_PS_excess", "UV_PS_excess", "UG_TPS_excess", "UV_TPS_excess"])

    with OUT_SUMMARY.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=summary_fields)
        writer.writeheader()
        for row in sorted(mine_rows, key=lambda r: r["MOF"]):
            writer.writerow({k: row.get(k, "") for k in summary_fields})

    abs_excess_fields = [
        "MOF", "basis", "PS_high_molkg", "PS_low_molkg", "TPS_low_molkg",
        "PS_high_mgg_reported", "PS_low_mgg_reported", "TPS_low_mgg_reported",
        "PS_high_mgg_corrected", "PS_low_mgg_corrected", "TPS_low_mgg_corrected",
        "PS_high_ratio_reported", "PS_low_ratio_reported", "TPS_low_ratio_reported",
        "PS_high_wt_pct", "PS_low_wt_pct", "TPS_low_wt_pct",
        "PS_high_uv_gL", "PS_low_uv_gL", "TPS_low_uv_gL",
        "UG_PS_wt_pct", "UV_PS_gL", "UG_TPS_wt_pct", "UV_TPS_gL",
    ]

    with OUT_ABS_EXCESS.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=abs_excess_fields)
        writer.writeheader()
        for row in sorted(mine_rows, key=lambda r: r["MOF"]):
            for basis in ["absolute", "excess"]:
                if basis == "excess" and not has_excess_any:
                    continue
                if basis == "absolute":
                    mol_key = "abs_molkg"
                    mgg_rep_key = "abs_mgg_reported"
                    mgg_corr_key = "abs_mgg_corrected"
                    ratio_key = "ratio_reported"
                    wt_key = "abs_wt_pct"
                    uv_key = "abs_uv_gL"
                else:
                    mol_key = "excess_molkg"
                    mgg_rep_key = "excess_mgg_reported"
                    mgg_corr_key = "excess_mgg_corrected"
                    ratio_key = "excess_ratio_reported"
                    wt_key = "excess_wt_pct"
                    uv_key = "excess_uv_gL"

                ps_h_m = to_float(row.get(f"PS_high_{mol_key}"))
                ps_l_m = to_float(row.get(f"PS_low_{mol_key}"))
                tp_l_m = to_float(row.get(f"TPS_low_{mol_key}"))
                # If this MOF has no excess values at all, skip excess row.
                if basis == "excess" and ps_h_m is None and ps_l_m is None and tp_l_m is None:
                    continue

                rec = {
                    "MOF": row["MOF"],
                    "basis": basis,
                    "PS_high_molkg": row.get(f"PS_high_{mol_key}", ""),
                    "PS_low_molkg": row.get(f"PS_low_{mol_key}", ""),
                    "TPS_low_molkg": row.get(f"TPS_low_{mol_key}", ""),
                    "PS_high_mgg_reported": row.get(f"PS_high_{mgg_rep_key}", ""),
                    "PS_low_mgg_reported": row.get(f"PS_low_{mgg_rep_key}", ""),
                    "TPS_low_mgg_reported": row.get(f"TPS_low_{mgg_rep_key}", ""),
                    "PS_high_mgg_corrected": row.get(f"PS_high_{mgg_corr_key}", ""),
                    "PS_low_mgg_corrected": row.get(f"PS_low_{mgg_corr_key}", ""),
                    "TPS_low_mgg_corrected": row.get(f"TPS_low_{mgg_corr_key}", ""),
                    "PS_high_ratio_reported": row.get(f"PS_high_{ratio_key}", ""),
                    "PS_low_ratio_reported": row.get(f"PS_low_{ratio_key}", ""),
                    "TPS_low_ratio_reported": row.get(f"TPS_low_{ratio_key}", ""),
                    "PS_high_wt_pct": row.get(f"PS_high_{wt_key}", ""),
                    "PS_low_wt_pct": row.get(f"PS_low_{wt_key}", ""),
                    "TPS_low_wt_pct": row.get(f"TPS_low_{wt_key}", ""),
                    "PS_high_uv_gL": row.get(f"PS_high_{uv_key}", ""),
                    "PS_low_uv_gL": row.get(f"PS_low_{uv_key}", ""),
                    "TPS_low_uv_gL": row.get(f"TPS_low_{uv_key}", ""),
                    "UG_PS_wt_pct": row.get("UG_PS" if basis == "absolute" else "UG_PS_excess", ""),
                    "UV_PS_gL": row.get("UV_PS" if basis == "absolute" else "UV_PS_excess", ""),
                    "UG_TPS_wt_pct": row.get("UG_TPS" if basis == "absolute" else "UG_TPS_excess", ""),
                    "UV_TPS_gL": row.get("UV_TPS" if basis == "absolute" else "UV_TPS_excess", ""),
                }
                writer.writerow(rec)

    # Dataset loading
    ps_path = DATASETS_DIR / "ps.csv"
    tps_path = DATASETS_DIR / "tps.csv"
    mof_db_path = None
    for p in sorted(DATASETS_DIR.glob("mof_db.*")):
        mof_db_path = p
        break

    ps_headers, ps_rows, ps_type = load_table(ps_path)
    tps_headers, tps_rows, tps_type = load_table(tps_path)
    if mof_db_path is None:
        raise RuntimeError("Could not find ~/datasets/mof_db.*")
    mof_headers, mof_rows, mof_type = load_table(mof_db_path)

    # Column detection
    ps_ref_col, ps_ref_cands = choose_best_column(ps_headers, include=["csd", "ref"], exclude=[])
    ps_name_col, ps_name_cands = choose_best_column(ps_headers, include=["name"], exclude=[])
    ps_ug_col, ps_ug_cands = choose_best_column(ps_headers, include=["ug", "ps"], exclude=["tps"])
    ps_uv_col, ps_uv_cands = choose_best_column(ps_headers, include=["uv", "ps"], exclude=["tps"])

    tps_ref_col, tps_ref_cands = choose_best_column(tps_headers, include=["csd", "ref"], exclude=[])
    tps_name_col, tps_name_cands = choose_best_column(tps_headers, include=["name"], exclude=[])
    tps_ug_col, tps_ug_cands = choose_best_column(tps_headers, include=["ug", "tps"], exclude=[])
    tps_uv_col, tps_uv_cands = choose_best_column(tps_headers, include=["uv", "tps"], exclude=[])

    mof_ref_col, mof_ref_cands = choose_best_column(mof_headers, include=["ref"], exclude=[])
    mof_name_col, mof_name_cands = choose_best_column(mof_headers, include=["name"], exclude=[])

    print(f"[DATASET] ps.csv type={ps_type} rows={len(ps_rows)}")
    print(f"[DATASET] tps.csv type={tps_type} rows={len(tps_rows)}")
    print(f"[DATASET] {mof_db_path.name} type={mof_type} rows={len(mof_rows)}")
    print(f"[COLUMNS][PS] ref={ps_ref_col} ug={ps_ug_col} uv={ps_uv_col} name={ps_name_col}")
    print(f"[COLUMNS][TPS] ref={tps_ref_col} ug={tps_ug_col} uv={tps_uv_col} name={tps_name_col}")
    print(f"[COLUMNS][MOF_DB] ref={mof_ref_col} name={mof_name_col}")
    print("[CANDIDATES][PS_UG]", ps_ug_cands[:5])
    print("[CANDIDATES][PS_UV]", ps_uv_cands[:5])
    print("[CANDIDATES][TPS_UG]", tps_ug_cands[:5])
    print("[CANDIDATES][TPS_UV]", tps_uv_cands[:5])

    # Excess hint from column/metadata names
    dataset_excess_hint = any("excess" in norm_header(h) for h in (ps_headers + tps_headers + mof_headers))

    ps_idx = build_dataset_index(ps_rows, [ps_ref_col, ps_name_col])
    tps_idx = build_dataset_index(tps_rows, [tps_ref_col, tps_name_col])
    mof_idx = build_dataset_index(mof_rows, [mof_ref_col, mof_name_col])

    def match_row(idx: Dict[str, Dict[str, str]], mine_mof: str) -> Tuple[Optional[Dict[str, str]], str]:
        for cand in id_candidates(mine_mof):
            if cand in idx:
                return idx[cand], cand
        # Try through MOF DB aliases
        for cand in id_candidates(mine_mof):
            mrow = mof_idx.get(cand)
            if not mrow:
                continue
            alias_vals = [mrow.get(mof_ref_col, ""), mrow.get(mof_name_col, "")]
            for av in alias_vals:
                for c2 in id_candidates(av):
                    if c2 in idx:
                        return idx[c2], c2
        return None, ""

    ps_join_fields = [
        "MOF", "MOF_norm", "dataset_match_key", "dataset_name", "dataset_refcode",
        "my_UG_PS_wt_pct", "my_UV_PS_g_L", "dataset_UG_PS_wt_pct", "dataset_UV_PS_g_L",
        "delta_UG_abs", "delta_UG_pct", "delta_UV_abs", "delta_UV_pct", "likely_mismatch_reason",
    ]
    tps_join_fields = [
        "MOF", "MOF_norm", "dataset_match_key", "dataset_name", "dataset_refcode",
        "my_UG_TPS_wt_pct", "my_UV_TPS_g_L", "dataset_UG_TPS_wt_pct", "dataset_UV_TPS_g_L",
        "delta_UG_abs", "delta_UG_pct", "delta_UV_abs", "delta_UV_pct", "likely_mismatch_reason",
    ]

    ps_join_rows = []
    tps_join_rows = []

    for row in sorted(mine_rows, key=lambda r: r["MOF"]):
        mof = row["MOF"]

        # PS join
        ps_row, ps_key = match_row(ps_idx, mof)
        my_ug_ps = to_float(row.get("UG_PS"))
        my_uv_ps = to_float(row.get("UV_PS"))
        ds_ug_ps = to_float(ps_row.get(ps_ug_col, "") if ps_row else "")
        ds_uv_ps = to_float(ps_row.get(ps_uv_col, "") if ps_row else "")
        ps_join_rows.append({
            "MOF": mof,
            "MOF_norm": norm_id(mof),
            "dataset_match_key": ps_key,
            "dataset_name": clean_text(ps_row.get(ps_name_col, "") if ps_row else ""),
            "dataset_refcode": clean_text(ps_row.get(ps_ref_col, "") if ps_row else ""),
            "my_UG_PS_wt_pct": fmt(my_ug_ps, 6),
            "my_UV_PS_g_L": fmt(my_uv_ps, 6),
            "dataset_UG_PS_wt_pct": fmt(ds_ug_ps, 6),
            "dataset_UV_PS_g_L": fmt(ds_uv_ps, 6),
            "delta_UG_abs": fmt(safe_sub(my_ug_ps, ds_ug_ps), 6),
            "delta_UG_pct": fmt(pct_diff(my_ug_ps, ds_ug_ps), 3),
            "delta_UV_abs": fmt(safe_sub(my_uv_ps, ds_uv_ps), 6),
            "delta_UV_pct": fmt(pct_diff(my_uv_ps, ds_uv_ps), 3),
            "likely_mismatch_reason": likely_reason(my_ug_ps, my_uv_ps, ds_ug_ps, ds_uv_ps, dataset_excess_hint),
        })

        # TPS join
        tps_row, tps_key = match_row(tps_idx, mof)
        my_ug_tps = to_float(row.get("UG_TPS"))
        my_uv_tps = to_float(row.get("UV_TPS"))
        ds_ug_tps = to_float(tps_row.get(tps_ug_col, "") if tps_row else "")
        ds_uv_tps = to_float(tps_row.get(tps_uv_col, "") if tps_row else "")
        tps_join_rows.append({
            "MOF": mof,
            "MOF_norm": norm_id(mof),
            "dataset_match_key": tps_key,
            "dataset_name": clean_text(tps_row.get(tps_name_col, "") if tps_row else ""),
            "dataset_refcode": clean_text(tps_row.get(tps_ref_col, "") if tps_row else ""),
            "my_UG_TPS_wt_pct": fmt(my_ug_tps, 6),
            "my_UV_TPS_g_L": fmt(my_uv_tps, 6),
            "dataset_UG_TPS_wt_pct": fmt(ds_ug_tps, 6),
            "dataset_UV_TPS_g_L": fmt(ds_uv_tps, 6),
            "delta_UG_abs": fmt(safe_sub(my_ug_tps, ds_ug_tps), 6),
            "delta_UG_pct": fmt(pct_diff(my_ug_tps, ds_ug_tps), 3),
            "delta_UV_abs": fmt(safe_sub(my_uv_tps, ds_uv_tps), 6),
            "delta_UV_pct": fmt(pct_diff(my_uv_tps, ds_uv_tps), 3),
            "likely_mismatch_reason": likely_reason(my_ug_tps, my_uv_tps, ds_ug_tps, ds_uv_tps, dataset_excess_hint),
        })

    with OUT_JOIN_PS.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=ps_join_fields)
        writer.writeheader()
        writer.writerows(ps_join_rows)

    with OUT_JOIN_TPS.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=tps_join_fields)
        writer.writeheader()
        writer.writerows(tps_join_rows)

    print(f"[WRITE] {OUT_SUMMARY}")
    print(f"[WRITE] {OUT_ABS_EXCESS}")
    print(f"[WRITE] {OUT_JOIN_PS}")
    print(f"[WRITE] {OUT_JOIN_TPS}")

    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        raise
